from io import BytesIO

from fastapi import UploadFile
from openpyxl import load_workbook


async def extract_company_rows_from_excel(upload: UploadFile) -> list[dict[str, str]]:
    payload = await upload.read()
    workbook = load_workbook(filename=BytesIO(payload), read_only=True, data_only=True)
    worksheet = workbook.active

    header_row_index = None
    company_col = None
    website_col = None

    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        normalized = [str(cell).strip().lower() if cell is not None else "" for cell in row]
        if "company name" in normalized and "website link" in normalized:
            header_row_index = row_index
            company_col = normalized.index("company name")
            website_col = normalized.index("website link")
            break

    if header_row_index is None or company_col is None or website_col is None:
        return []

    companies: list[dict[str, str]] = []
    for row in worksheet.iter_rows(min_row=header_row_index + 1, values_only=True):
        company_name = str(row[company_col]).strip() if len(row) > company_col and row[company_col] else ""
        website_link = str(row[website_col]).strip() if len(row) > website_col and row[website_col] else ""
        if not company_name or not website_link:
            continue
        companies.append({"company_name": company_name, "website_link": website_link})

    return companies

